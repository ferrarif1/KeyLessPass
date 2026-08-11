#!/usr/bin/env python3
"""Convert the public SOUPS 2022 PCP workbook into auditable PolicySpec rows.

Requires openpyxl. Unsupported semantics are rejected instead of approximated.
Every exact translation is emitted for compilation under the evaluator's
separate state, memory, and wall-time budgets.
"""

import argparse
import hashlib
import json
from pathlib import Path

from openpyxl import load_workbook


LOWER = "abcdefghijklmnopqrstuvwxyz"
UPPER = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
DIGITS = "0123456789"
SYMBOLS = "!\"#$%&'()*+,-./:;<=>?@[\\]^_`{|}~ "
CHARSETS = {
    "lower": LOWER,
    "upper": UPPER,
    "alphabet": LOWER + UPPER,
    "digits": DIGITS,
    "symbols": SYMBOLS,
}


def reject(reason):
    return {"translationStatus": "rejected", "reason": reason}


def translate(policy, exclusions):
    if exclusions:
        return reject("context_dependent_policy_exclusions")
    for key, reason in (
        ("rules", "disjunctive_rules"),
        ("require_subset", "subset_of_classes_constraint"),
        ("charsets", "custom_charsets"),
        ("charset_requirements", "per_charset_location_or_run_constraint"),
    ):
        if key in policy:
            return reject(reason)
    if policy.get("max_length") is None:
        return reject("unbounded_max_length")
    if policy["max_length"] > 128:
        return reject("protocol_max_length_128")

    required = policy.get("require", [])
    unknown = sorted(set(required) - set(CHARSETS))
    if unknown:
        return reject("unknown_character_class:" + ",".join(unknown))

    spec = {
        "policyIrVersion": 1,
        "minLength": policy.get("min_length", 1),
        "maxLength": policy["max_length"],
        "alphabet": LOWER + UPPER + DIGITS + SYMBOLS,
        "forbiddenCharacters": "",
        "classes": [
            {"name": name, "alphabet": CHARSETS[name], "minCount": 1, "maxCount": None}
            for name in required
        ],
        "fixedCharacters": [],
        "fixedPrefix": "",
        "fixedSuffix": "",
        "forbiddenFirstCharacters": "",
        "forbiddenLastCharacters": "",
        "maxTotalPerCharacter": None,
        "maxIdenticalRun": policy.get("max_consecutive"),
        "maxSequentialRun": None,
        "forbiddenSubstrings": policy.get("prohibited_substrings", []),
    }
    return {
        "translationStatus": "translated",
        "policySpec": spec,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("workbook", type=Path)
    parser.add_argument("output", type=Path)
    args = parser.parse_args()

    workbook_bytes = args.workbook.read_bytes()
    book = load_workbook(args.workbook, read_only=True, data_only=True)
    sheet = book["policy_websites"]
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value) if value is not None else "_index" for value in next(rows)]

    records = []
    for source_row, values in enumerate(rows, start=2):
        row = dict(zip(headers, values))
        if row.get("policy") is None:
            continue
        policy = json.loads(row["policy"])
        translated = translate(policy, row.get("policy_exclusions"))
        records.append({
            "sourceRow": source_row,
            "website": row["website"],
            "globalRank": row.get("global_rank"),
            "sourcePolicy": policy,
            **translated,
        })

    payload = {
        "schemaVersion": 1,
        "source": {
            "title": "Improving Password Generation Through the Design of a Password Composition Policy Description Language",
            "authors": "Gautam, Lalani, and Ruoti",
            "venue": "SOUPS 2022",
            "archiveUrl": "https://userlab.utk.edu/files/data/ruoti/2022/gautam2022improving.zip",
            "workbookPath": "pcp dataset/clean_data.xlsx",
            "workbookSha256": hashlib.sha256(workbook_bytes).hexdigest(),
            "datasetScope": "historical website policies collected by the source authors; not current live policies",
        },
        "translation": {
            "tool": "experiments/scripts/import_soups2022_pcp.py",
            "unsupportedSemanticsAreRejected": True,
            "compilationPrefilter": None,
        },
        "records": records,
    }
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(json.dumps(payload, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
